import os
import io
import sqlite3
from datetime import datetime
import requests
from flask import Flask, request, jsonify, send_file
import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)

VERIFY_TOKEN = os.environ.get("VERIFY_TOKEN", "my_secret_token")
WHATSAPP_TOKEN = os.environ.get("WHATSAPP_TOKEN")
PHONE_NUMBER_ID = os.environ.get("PHONE_NUMBER_ID")

DB_PATH = os.environ.get("DB_PATH", "yuk_novbe.db")

# ⚠️ Bura Cavid, Xəlil, Azər-in WhatsApp nömrələrini yazın
# Format: ölkə kodu + nömrə, "+" işarəsi OLMADAN (məs. "994501234567")
STAFF = {
    "994XXXXXXXXX": "Cavid",
    "994XXXXXXXXX": "Xəlil",
    "994XXXXXXXXX": "Azər",
}

DEPO_CATEGORIES = {
    "1": "Adsız palet",
    "2": "Silinmə",
    "3": "GZ Zone",
}

# ---------- DB ----------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_number TEXT,
            category TEXT,
            subcategory TEXT,
            full_name TEXT,
            photo_media_id TEXT,
            note TEXT,
            status TEXT DEFAULT 'Gözləyir',
            assigned_staff TEXT,
            created_at TEXT,
            closed_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS user_state (
            phone_number TEXT PRIMARY KEY,
            step TEXT,
            category TEXT,
            subcategory TEXT,
            full_name TEXT,
            photo_media_id TEXT
        )
    """)
    conn.commit()
    conn.close()


init_db()

# ---------- State helpers ----------

def get_state(phone):
    conn = get_db()
    row = conn.execute("SELECT * FROM user_state WHERE phone_number=?", (phone,)).fetchone()
    conn.close()
    return dict(row) if row else None


def set_state(phone, **kwargs):
    conn = get_db()
    existing = conn.execute("SELECT * FROM user_state WHERE phone_number=?", (phone,)).fetchone()
    if existing:
        fields = ", ".join(f"{k}=?" for k in kwargs)
        conn.execute(f"UPDATE user_state SET {fields} WHERE phone_number=?", (*kwargs.values(), phone))
    else:
        cols = ["phone_number"] + list(kwargs.keys())
        placeholders = ", ".join("?" * len(cols))
        conn.execute(f"INSERT INTO user_state ({', '.join(cols)}) VALUES ({placeholders})", (phone, *kwargs.values()))
    conn.commit()
    conn.close()


def clear_state(phone):
    conn = get_db()
    conn.execute("DELETE FROM user_state WHERE phone_number=?", (phone,))
    conn.commit()
    conn.close()

# ---------- Ticket helpers ----------

def create_ticket(customer_number, category, subcategory, full_name, photo_media_id, note):
    conn = get_db()
    cur = conn.execute("""
        INSERT INTO tickets (customer_number, category, subcategory, full_name, photo_media_id, note, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 'Gözləyir', ?)
    """, (customer_number, category, subcategory, full_name, photo_media_id, note, datetime.now().isoformat()))
    conn.commit()
    ticket_id = cur.lastrowid
    conn.close()
    return ticket_id


def get_pending_tickets():
    conn = get_db()
    rows = conn.execute("SELECT * FROM tickets WHERE status='Gözləyir' ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_ticket(ticket_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM tickets WHERE id=?", (ticket_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def assign_ticket(ticket_id, staff_number):
    conn = get_db()
    conn.execute("UPDATE tickets SET status='Araşdırılır', assigned_staff=? WHERE id=?", (staff_number, ticket_id))
    conn.commit()
    conn.close()


def close_ticket(ticket_id):
    conn = get_db()
    conn.execute("UPDATE tickets SET status='Bitdi', closed_at=? WHERE id=?", (datetime.now().isoformat(), ticket_id))
    conn.commit()
    conn.close()


def get_staff_active_ticket(staff_number):
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM tickets WHERE assigned_staff=? AND status='Araşdırılır'",
        (staff_number,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None

# ---------- WhatsApp helpers ----------

def send_whatsapp_message(to_number, text):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {"Authorization": f"Bearer {WHATSAPP_TOKEN}", "Content-Type": "application/json"}
    payload = {"messaging_product": "whatsapp", "to": to_number, "type": "text", "text": {"body": text}}
    r = requests.post(url, json=payload, headers=headers)
    print(f"[send_text] {r.status_code} {r.text}")
    return r


def send_whatsapp_image(to_number, media_id, caption=""):
    url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/messages"
    headers = {"Authorization": f"Bearer {WHATSAPP_TOKEN}", "Content-Type": "application/json"}
    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "image",
        "image": {"id": media_id, "caption": caption},
    }
    r = requests.post(url, json=payload, headers=headers)
    print(f"[send_image] {r.status_code} {r.text}")
    return r


def download_and_reupload_media(media_id):
    """Bir istifadəçidən gələn şəkli başqa nömrəyə göndərmək üçün yenidən yükləyir."""
    headers = {"Authorization": f"Bearer {WHATSAPP_TOKEN}"}
    meta_url = f"https://graph.facebook.com/v18.0/{media_id}"
    meta_resp = requests.get(meta_url, headers=headers).json()
    file_url = meta_resp.get("url")
    if not file_url:
        return None
    file_resp = requests.get(file_url, headers=headers)
    files = {"file": ("photo.jpg", file_resp.content, meta_resp.get("mime_type", "image/jpeg"))}
    data = {"messaging_product": "whatsapp"}
    upload_url = f"https://graph.facebook.com/v18.0/{PHONE_NUMBER_ID}/media"
    upload_resp = requests.post(upload_url, headers=headers, data=data, files=files).json()
    return upload_resp.get("id")

# ---------- Webhook ----------

@app.route("/webhook", methods=["GET"])
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    return "Forbidden", 403


@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.get_json()

    if data.get("object") != "whatsapp_business_account":
        return jsonify({"status": "not a whatsapp event"}), 404

    for entry in data.get("entry", []):
        for change in entry.get("changes", []):
            value = change.get("value", {})
            messages = value.get("messages", [])
            if not messages:
                continue

            msg = messages[0]
            from_number = msg.get("from")
            msg_type = msg.get("type")

            if from_number in STAFF:
                handle_staff_message(from_number, msg, msg_type)
            else:
                handle_customer_message(from_number, msg, msg_type)

    return jsonify({"status": "success"}), 200

# ---------- Müştəri axını ----------

MAIN_MENU = "Xoş gəlmisiniz!\n1 - Depo\n2 - Nəqliyyat\n\nCavab olaraq rəqəm yazın."
DEPO_MENU = "Depo bölməsi:\n1 - Adsız palet\n2 - Silinmə\n3 - GZ Zone"
NEQLIYYAT_MENU = "Nəqliyyat bölməsi:\n1 - Sürücü\n2 - Dispetçer"


def handle_customer_message(from_number, msg, msg_type):
    state = get_state(from_number)
    text_body = msg.get("text", {}).get("body", "").strip() if msg_type == "text" else ""

    if text_body.lower() in ["menyu", "menu", "başla", "salam"]:
        clear_state(from_number)
        send_whatsapp_message(from_number, MAIN_MENU)
        return

    if not state:
        if text_body == "1":
            set_state(from_number, step="depo_menu")
            send_whatsapp_message(from_number, DEPO_MENU)
        elif text_body == "2":
            set_state(from_number, step="nəqliyyat_menu")
            send_whatsapp_message(from_number, NEQLIYYAT_MENU)
        else:
            send_whatsapp_message(from_number, MAIN_MENU)
        return

    step = state["step"]

    if step == "depo_menu":
        if text_body in DEPO_CATEGORIES:
            set_state(from_number, step="ask_name", category="Depo", subcategory=DEPO_CATEGORIES[text_body])
            send_whatsapp_message(from_number, "Zəhmət olmasa ad və soyadınızı yazın:")
        else:
            send_whatsapp_message(from_number, DEPO_MENU)
        return

    if step == "nəqliyyat_menu":
        if text_body == "1":
            send_whatsapp_message(from_number, "Sürücü bölməsi tezliklə əlavə olunacaq.")
        elif text_body == "2":
            send_whatsapp_message(from_number, "Dispetçer bölməsi tezliklə əlavə olunacaq.")
        else:
            send_whatsapp_message(from_number, NEQLIYYAT_MENU)
            return
        clear_state(from_number)
        return

    if step == "ask_name":
        if text_body:
            set_state(from_number, step="ask_photo", full_name=text_body)
            send_whatsapp_message(from_number, "Təşəkkürlər. İndi şəklini göndərin:")
        else:
            send_whatsapp_message(from_number, "Zəhmət olmasa ad və soyadınızı mətn kimi yazın.")
        return

    if step == "ask_photo":
        if msg_type == "image":
            media_id = msg.get("image", {}).get("id")
            set_state(from_number, step="ask_note", photo_media_id=media_id)
            send_whatsapp_message(from_number, "Şəkil qəbul edildi. İndi qeydinizi yazın:")
        else:
            send_whatsapp_message(from_number, "Zəhmət olmasa şəkil göndərin.")
        return

    if step == "ask_note":
        if text_body:
            ticket_id = create_ticket(
                customer_number=from_number,
                category=state["category"],
                subcategory=state["subcategory"],
                full_name=state["full_name"],
                photo_media_id=state["photo_media_id"],
                note=text_body,
            )
            clear_state(from_number)
            send_whatsapp_message(
                from_number,
                f"Sorğunuz qeydə alındı. Növbə nömrəniz: #{ticket_id}\nStatus: Gözləyir"
            )
            notify_staff_new_ticket(ticket_id)
        else:
            send_whatsapp_message(from_number, "Zəhmət olmasa qeydinizi mətn kimi yazın.")
        return


def notify_staff_new_ticket(ticket_id):
    ticket = get_ticket(ticket_id)
    text = (
        f"🆕 Yeni sorğu #{ticket_id}\n"
        f"Bölmə: {ticket['category']} - {ticket['subcategory']}\n"
        f"Ad: {ticket['full_name']}\n"
        f"Qeyd: {ticket['note']}\n\n"
        f"Götürmək üçün: götür {ticket_id}"
    )
    for staff_number in STAFF:
        send_whatsapp_message(staff_number, text)

# ---------- Staff axını ----------

def handle_staff_message(staff_number, msg, msg_type):
    text_body = msg.get("text", {}).get("body", "").strip() if msg_type == "text" else ""
    text_lower = text_body.lower()

    active_ticket = get_staff_active_ticket(staff_number)

    if text_lower in ["sıra", "növbə"]:
        pending = get_pending_tickets()
        if not pending:
            send_whatsapp_message(staff_number, "Növbədə gözləyən sorğu yoxdur.")
        else:
            lines = [f"#{t['id']} - {t['subcategory']} - {t['full_name']}" for t in pending]
            send_whatsapp_message(staff_number, "Gözləyən sorğular:\n" + "\n".join(lines))
        return

    if text_lower.startswith("götür"):
        parts = text_body.split()
        if len(parts) == 2 and parts[1].isdigit():
            ticket_id = int(parts[1])
            ticket = get_ticket(ticket_id)
            if not ticket or ticket["status"] != "Gözləyir":
                send_whatsapp_message(staff_number, "Bu sorğu artıq götürülüb və ya mövcud deyil.")
                return
            if active_ticket:
                send_whatsapp_message(
                    staff_number,
                    f"Əvvəlcə #{active_ticket['id']} sorğusunu 'bitdi' yazaraq bağlayın."
                )
                return
            assign_ticket(ticket_id, staff_number)
            send_whatsapp_message(
                staff_number,
                f"#{ticket_id} sizə təyin edildi. Cavab yazmaq üçün mesaj yazın, bitirdikdə 'bitdi' yazın."
            )
            if ticket["photo_media_id"]:
                new_media_id = download_and_reupload_media(ticket["photo_media_id"])
                if new_media_id:
                    send_whatsapp_image(staff_number, new_media_id, caption=f"#{ticket_id} - {ticket['full_name']}")
            send_whatsapp_message(
                ticket["customer_number"],
                f"Sorğunuz (#{ticket_id}) araşdırılır. Zəhmət olmasa gözləyin."
            )
        else:
            send_whatsapp_message(staff_number, "Format: götür <sorğu nömrəsi>  (məs. götür 5)")
        return

    if text_lower == "bitdi":
        if active_ticket:
            close_ticket(active_ticket["id"])
            send_whatsapp_message(staff_number, f"#{active_ticket['id']} bağlandı. Yeni sorğu üçün 'sıra' yazın.")
        else:
            send_whatsapp_message(staff_number, "Aktiv sorğunuz yoxdur.")
        return

    # Aktiv sorğu varsa, mesaj birbaşa müştəriyə yönləndirilir
    if active_ticket:
        if msg_type == "text":
            send_whatsapp_message(active_ticket["customer_number"], text_body)
        elif msg_type == "image":
            media_id = msg.get("image", {}).get("id")
            new_media_id = download_and_reupload_media(media_id)
            if new_media_id:
                send_whatsapp_image(active_ticket["customer_number"], new_media_id)
        return

    send_whatsapp_message(staff_number, "Aktiv sorğunuz yoxdur. Gözləyən sorğulara baxmaq üçün 'sıra' yazın.")

# ---------- Excel export ----------

@app.route("/export", methods=["GET"])
def export_excel():
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    rows = conn.execute("SELECT * FROM tickets WHERE date(created_at) = ?", (today,)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sorgular {today}"
    headers = ["ID", "Müştəri nömrəsi", "Bölmə", "Alt bölmə", "Ad Soyad", "Qeyd",
               "Status", "Cavabdeh", "Yaradılma tarixi", "Bağlanma tarixi"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["id"], row["customer_number"], row["category"], row["subcategory"],
            row["full_name"], row["note"], row["status"],
            STAFF.get(row["assigned_staff"], row["assigned_staff"] or ""),
            row["created_at"], row["closed_at"] or ""
        ])

    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(col) + 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sorgular_{today}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
